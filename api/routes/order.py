"""
Ordering slice — warehouse picklist, production sheets, service-dial.

Endpoints (prefix /order):
  GET /order/picklist?date          — network-wide picklist (product roll-up)
  GET /order/picklist.csv?date      — same, StreamingResponse CSV
  GET /order/production/{id}?date   — per-outlet breakdown for one product
  GET /order/by-warehouse?date      — picklist split per supplying warehouse
  GET /order/dial?date              — service-vs-waste sweep for dial chart

All queries use the neutral DuckDB views from deps/duck.py.
Warehouse mapping via deps/order_views.py (never touches duck.py).
"""
from __future__ import annotations
import io, csv, pathlib
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse, JSONResponse
from deps.duck import q, con
from deps.order_views import ensure_views

import numpy as np

router = APIRouter(prefix="/order", tags=["ordering"])

_ROOT = pathlib.Path(__file__).resolve().parent.parent.parent
_ORDER_PLAN = _ROOT / "data" / "predictions" / "order_plan.parquet"
_DIM_BRANCH = _ROOT / "data" / "raw" / "dim_branch.parquet"
_DIM_PRODUCT = _ROOT / "data" / "raw" / "dim_product.parquet"

# ── ensure dim_warehouse view is available when this module loads ──
ensure_views()


# ── helpers ────────────────────────────────────────────────────────

def _latest_date() -> str:
    r = q("SELECT max(date) d FROM forecast")
    return str(r[0]["d"]) if r and r[0]["d"] else ""


def _q_at(cr: float, p50: np.ndarray, p85: np.ndarray, p95: np.ndarray) -> np.ndarray:
    """
    Piecewise-linear demand quantile at critical ratio cr (scalar), given P50/P85/P95 arrays.
    Reimplemented here so this route has zero dependency on src/order_qty.py.
    """
    cr_arr = np.full(len(p50), cr, dtype=float)
    s1 = p50 * (cr_arr / 0.5)
    s2 = p50 + (cr_arr - 0.5) / 0.35 * (p85 - p50)
    s3 = p85 + (cr_arr - 0.85) / 0.10 * (p95 - p85)
    s4 = p95
    out = np.select(
        [cr_arr < 0.5, cr_arr < 0.85, cr_arr < 0.95],
        [s1, s2, s3],
        default=s4,
    )
    return np.clip(np.round(out), 0, None)


def _sim(demand: np.ndarray, order: np.ndarray,
         cu: np.ndarray, co: np.ndarray) -> dict:
    """Simulate one service-level scenario."""
    under = np.maximum(demand - order, 0.0)
    over  = np.maximum(order  - demand, 0.0)
    cost  = float((cu * under + co * over).sum())
    total_order  = float(order.sum())
    total_demand = float(demand.sum())
    stockout_pct = float((demand > order).mean() * 100)
    waste_pct    = float(over.sum() / max(total_order, 1) * 100)
    fill_pct     = float(np.minimum(order, demand).sum() / max(total_demand, 1) * 100)
    return dict(
        cost=round(cost, 0),
        stockout_pct=round(stockout_pct, 1),
        waste_pct=round(waste_pct, 1),
        fill_pct=round(fill_pct, 1),
    )


# ── Excel export: order plan by OUTLET × SKU ────────────────────────
def _export_rows_fabric(date: str | None):
    """Order-plan rows from the Fabric cfc_order_plan table, joined to branch/product
    masters (price ← Ref_ProductMaster.ListPrice; abc not stored in Fabric → blank).
    Returns rows shaped like the local DuckDB query, or None when Fabric is off/fails
    (→ caller falls back to local parquet). Never raises into the request."""
    from deps import fabric
    if not fabric.enabled():
        return None
    try:
        OP, B, P = fabric.table("cfc_order_plan"), fabric.table("Ref_BranchMaster"), fabric.table("Ref_ProductMaster")
        if not date:
            mx = fabric.q(f"SELECT MAX(date) mx FROM {OP}")
            if not mx or mx[0]["mx"] is None:
                return None
            date = str(mx[0]["mx"])[:10]
        rows = fabric.q(
            f"SELECT b.BranchName outlet, b.BranchCode outlet_code, p.ProductCode sku, "
            f"p.ProductName product, CAST(NULL AS varchar) abc, o.p50, o.p85, o.p95, "
            f"o.order_qty, p.ListPrice price "
            f"FROM {OP} o "
            f"LEFT JOIN {B} b ON o.BranchId = b.BranchId "
            f"LEFT JOIN {P} p ON o.ProductId = p.ProductId "
            f"WHERE CAST(o.date AS DATE) = ? "
            f"ORDER BY b.BranchName, o.order_qty DESC", (date,))
        # normalise to positional tuples matching the local query column order;
        # cast Fabric Decimals→float so downstream float math (_qty_at) works.
        def _n(v): return None if v is None else float(v)
        return [(r["outlet"], r["outlet_code"], r["sku"], r["product"], r["abc"],
                 _n(r["p50"]), _n(r["p85"]), _n(r["p95"]), _n(r["order_qty"]), _n(r["price"])) for r in rows]
    except Exception:
        return None


@router.get("/export.xlsx")
def export_xlsx(
    date: str | None = Query(None, description="YYYY-MM-DD; default = latest plan date"),
    service_level: int | None = Query(None, ge=30, le=95,
                                      description="re-pick order qty at this service level %"),
):
    """One row per outlet × SKU for a given day (from the Fabric cfc_order_plan table
    when USE_FABRIC=1, else local order_plan.parquet), joined to outlet + product names.
    Sheet 1 = flat plan, Sheet 2 = per-outlet summary."""
    rows = _export_rows_fabric(date)          # None → Fabric off/failed → local fallback
    if rows is None:
        if not _ORDER_PLAN.exists():
            return JSONResponse({"error": "no order plan built yet"}, status_code=404)
        import duckdb
        c = duckdb.connect()
        op, db, dp = _ORDER_PLAN.as_posix(), _DIM_BRANCH.as_posix(), _DIM_PRODUCT.as_posix()
        if not date:
            date = str(c.execute(f"SELECT max(date) FROM read_parquet('{op}')").fetchone()[0])[:10]
        rows = c.execute(f"""
            SELECT b.BranchName   AS outlet,
                   b.BranchCode   AS outlet_code,
                   p.ProductCode  AS sku,
                   p.ProductName  AS product,
                   o.abc          AS abc,
                   o.p50, o.p85, o.p95,
                   o.order_qty,
                   o.price
            FROM read_parquet('{op}') o
            LEFT JOIN read_parquet('{db}') b ON o.BranchId = b.BranchId
            LEFT JOIN read_parquet('{dp}') p ON o.ProductId = p.ProductId
            WHERE CAST(o.date AS DATE) = DATE '{date}'
            ORDER BY b.BranchName, o.order_qty DESC
        """).fetchall()
        c.close()

    def _qty_at(p50, p85, p95, sl):
        cr = sl / 100.0
        if cr < 0.5:   v = p50 * (cr / 0.5)
        elif cr < 0.85: v = p50 + (cr - 0.5) / 0.35 * (p85 - p50)
        elif cr < 0.95: v = p85 + (cr - 0.85) / 0.10 * (p95 - p85)
        else:          v = p95
        return max(round(v), 0)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Order plan"
    headers = ["Outlet", "Outlet code", "SKU code", "Product", "ABC",
               "P50", "P85", "P95", "Order qty", "Est. value (₭)"]
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="232834")
    hfont = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal="left")

    per_outlet: dict[str, list] = {}
    for r in rows:
        outlet, ocode, sku, product, abc, p50, p85, p95, oqty, price = r
        qty = _qty_at(p50 or 0, p85 or 0, p95 or 0, service_level) if service_level else (oqty or 0)
        val = round((qty or 0) * (price or 0))
        ws.append([outlet, ocode, sku, product, abc,
                   round(p50 or 0, 1), round(p85 or 0, 1), round(p95 or 0, 1), qty, val])
        agg = per_outlet.setdefault(outlet or "—", [0, 0])
        agg[0] += qty; agg[1] += val

    # summary sheet
    ws2 = wb.create_sheet("Summary by outlet")
    ws2.append(["Outlet", "Total order qty", "Est. value (₭)"])
    for col in range(1, 4):
        cc = ws2.cell(row=1, column=col); cc.fill = hfill; cc.font = hfont
    for outlet, (q_, v_) in sorted(per_outlet.items(), key=lambda x: -x[1][1]):
        ws2.append([outlet, q_, v_])

    for w in (ws, ws2):
        for column_cells in w.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells) + 2
            w.column_dimensions[column_cells[0].column_letter].width = min(width, 40)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"order_plan_{date}{'_sl' + str(service_level) if service_level else ''}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── picklist (JSON) ─────────────────────────────────────────────────

@router.get("/picklist")
def picklist(date: str | None = Query(None)):
    date = date or _latest_date()
    rows = q("""
        SELECT
            f.product_id,
            COALESCE(p.product_name, CAST(f.product_id AS VARCHAR)) AS product_name,
            p.category,
            ROUND(SUM(f.order_qty), 0)          AS order_units,
            COUNT(DISTINCT f.outlet_id)          AS outlets,
            ROUND(SUM(f.order_qty * f.price), 0) AS value_ks
        FROM forecast f
        LEFT JOIN dim_product p USING (product_id)
        WHERE f.date = ?
        GROUP BY f.product_id, p.product_name, p.category
        ORDER BY order_units DESC
    """, [date])

    total_units = sum(r["order_units"] or 0 for r in rows)
    total_value = sum(r["value_ks"]   or 0 for r in rows)
    return {
        "date": date,
        "totals": {
            "products": len(rows),
            "order_units": round(total_units, 0),
            "value_ks": round(total_value, 0),
        },
        "rows": rows,
    }


# ── picklist (CSV download) ─────────────────────────────────────────

@router.get("/picklist.csv")
def picklist_csv(date: str | None = Query(None)):
    date = date or _latest_date()
    rows = q("""
        SELECT
            f.product_id,
            COALESCE(p.product_code, '') AS product_code,
            COALESCE(p.product_name, CAST(f.product_id AS VARCHAR)) AS product_name,
            COALESCE(p.category, '')     AS category,
            ROUND(SUM(f.order_qty), 0)          AS order_units,
            COUNT(DISTINCT f.outlet_id)          AS outlets,
            ROUND(SUM(f.order_qty * f.price), 0) AS value_ks
        FROM forecast f
        LEFT JOIN dim_product p USING (product_id)
        WHERE f.date = ?
        GROUP BY f.product_id, p.product_code, p.product_name, p.category
        ORDER BY order_units DESC
    """, [date])

    buf = io.StringIO()
    writer = csv.DictWriter(
        buf,
        fieldnames=["product_id", "product_code", "product_name", "category",
                    "order_units", "outlets", "value_ks"],
    )
    writer.writeheader()
    writer.writerows(rows)
    buf.seek(0)

    filename = f"picklist_{date}.csv"
    return StreamingResponse(
        iter([buf.read()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── production sheet (one product across all outlets) ───────────────

@router.get("/production/{product_id}")
def production(product_id: int, date: str | None = Query(None)):
    date = date or _latest_date()

    # product header
    head = q("""
        SELECT p.product_name, p.category, p.product_code
        FROM dim_product p WHERE p.product_id = ?
    """, [product_id])
    product_name = head[0]["product_name"] if head else str(product_id)
    category     = head[0]["category"]     if head else None
    product_code = head[0]["product_code"] if head else None

    rows = q("""
        SELECT
            f.outlet_id,
            COALESCE(b.outlet_name, CAST(f.outlet_id AS VARCHAR)) AS outlet_name,
            b.brand,
            ROUND(f.order_qty, 0)   AS order_qty,
            ROUND(f.expected,  1)   AS expected,
            ROUND(f.safe,      1)   AS safe,
            ROUND(f.actual,    1)   AS actual,
            CASE
                WHEN f.expected > COALESCE(f.actual, f.expected) * 1.05 THEN 'up'
                WHEN f.expected < COALESCE(f.actual, f.expected) * 0.95 THEN 'down'
                ELSE 'flat'
            END AS trend
        FROM forecast f
        LEFT JOIN dim_branch b USING (outlet_id)
        WHERE f.product_id = ? AND f.date = ?
        ORDER BY f.order_qty DESC
    """, [product_id, date])

    make_total = sum(r["order_qty"] or 0 for r in rows)
    return {
        "date": date,
        "product_id": product_id,
        "product_name": product_name,
        "product_code": product_code,
        "category": category,
        "make_total": round(make_total, 0),
        "outlets_count": len(rows),
        "rows": rows,
    }


# ── by-warehouse split ───────────────────────────────────────────────

@router.get("/by-warehouse")
def by_warehouse(date: str | None = Query(None)):
    """
    Picklist split per supplying warehouse.
    Mapping: dim_branch.StockOutId -> dim_warehouse.WarehouseId.
    Coverage is thin in the current dim (only ~1/77 branches resolve).
    When <10% map, returns a single network-wide bucket with a note.
    """
    ensure_views()
    date = date or _latest_date()

    # check mapping coverage
    coverage = q("""
        SELECT
            COUNT(DISTINCT b.outlet_id)                                          AS total_branches,
            COUNT(DISTINCT CASE WHEN w.warehouse_id IS NOT NULL
                           THEN b.outlet_id END)                                 AS mapped_branches
        FROM dim_branch b
        LEFT JOIN dim_warehouse w ON w.warehouse_id = CAST(b.outlet_id AS INTEGER)
        WHERE b.outlet_id IN (SELECT DISTINCT outlet_id FROM forecast WHERE date = ?)
    """, [date])

    total   = coverage[0]["total_branches"] if coverage else 0
    mapped  = coverage[0]["mapped_branches"] if coverage else 0
    sparse  = total == 0 or (mapped / total) < 0.10

    if sparse:
        # network-wide fallback
        rows = q("""
            SELECT
                f.product_id,
                COALESCE(p.product_name, CAST(f.product_id AS VARCHAR)) AS product_name,
                ROUND(SUM(f.order_qty), 0)          AS order_units,
                COUNT(DISTINCT f.outlet_id)          AS outlets,
                ROUND(SUM(f.order_qty * f.price), 0) AS value_ks
            FROM forecast f
            LEFT JOIN dim_product p USING (product_id)
            WHERE f.date = ?
            GROUP BY f.product_id, p.product_name
            ORDER BY order_units DESC
        """, [date])

        return {
            "date": date,
            "note": (
                f"Warehouse mapping is sparse ({mapped}/{total} branches resolved). "
                "Showing network-wide totals. Populate StockOutId in dim_branch to enable per-warehouse splits."
            ),
            "warehouses": [
                {
                    "warehouse_id": None,
                    "warehouse_name": "Network-wide",
                    "outlets_count": total,
                    "order_units": round(sum(r["order_units"] or 0 for r in rows), 0),
                    "value_ks": round(sum(r["value_ks"] or 0 for r in rows), 0),
                    "rows": rows,
                }
            ],
        }

    # full per-warehouse breakdown (for future when mapping is populated)
    rows = q("""
        SELECT
            COALESCE(w.warehouse_id, -1)           AS warehouse_id,
            COALESCE(w.warehouse_name, 'Unmapped') AS warehouse_name,
            f.product_id,
            COALESCE(p.product_name, CAST(f.product_id AS VARCHAR)) AS product_name,
            ROUND(SUM(f.order_qty), 0)          AS order_units,
            COUNT(DISTINCT f.outlet_id)          AS outlets,
            ROUND(SUM(f.order_qty * f.price), 0) AS value_ks
        FROM forecast f
        LEFT JOIN dim_branch  b USING (outlet_id)
        LEFT JOIN dim_warehouse w ON w.warehouse_id = CAST(b.outlet_id AS INTEGER)
        LEFT JOIN dim_product  p USING (product_id)
        WHERE f.date = ?
        GROUP BY w.warehouse_id, w.warehouse_name, f.product_id, p.product_name
        ORDER BY warehouse_id, order_units DESC
    """, [date])

    # group by warehouse
    warehouses: dict[int, dict] = {}
    for r in rows:
        wid  = r["warehouse_id"]
        wnam = r["warehouse_name"]
        if wid not in warehouses:
            warehouses[wid] = dict(
                warehouse_id=wid, warehouse_name=wnam,
                order_units=0, value_ks=0, outlets_count=0, rows=[],
            )
        wh = warehouses[wid]
        wh["rows"].append({k: v for k, v in r.items()
                           if k not in ("warehouse_id", "warehouse_name")})
        wh["order_units"] += r["order_units"] or 0
        wh["value_ks"]    += r["value_ks"]    or 0
        wh["outlets_count"] = r["outlets"] or 0

    return {
        "date": date,
        "note": None,
        "warehouses": list(warehouses.values()),
    }


# ── service dial (sweep) ─────────────────────────────────────────────

@router.get("/dial")
def dial(date: str | None = Query(None)):
    """
    Service-vs-waste sweep for the dial chart.
    Pulls all (expected, safe, max_safe, actual, price) for the date.
    Sweeps service levels [30, 50, 70, 85, 95].
    Uses flat GM=0.35, full same-day spoilage (Co = price*(1-GM)).
    Returns points suitable for a dual-axis chart (cost line + bars).
    """
    date = date or _latest_date()
    GM   = 0.35

    raw = q("""
        SELECT
            f.expected   AS p50,
            f.safe       AS p85,
            f.max_safe   AS p95,
            f.actual     AS demand,
            f.price      AS price
        FROM forecast f
        WHERE f.date = ?
    """, [date])

    if not raw:
        return {"date": date, "points": []}

    p50    = np.array([r["p50"]    or 0.0 for r in raw], dtype=float)
    p85    = np.array([r["p85"]    or 0.0 for r in raw], dtype=float)
    p95    = np.array([r["p95"]    or 0.0 for r in raw], dtype=float)
    demand = np.array([r["demand"] or 0.0 for r in raw], dtype=float)
    price  = np.array([r["price"]  or 0.0 for r in raw], dtype=float)

    cu = price * GM
    co = price * (1.0 - GM)   # full same-day spoilage, no salvage

    SERVICE_LEVELS = [30, 50, 70, 85, 95]
    points = []
    for sl in SERVICE_LEVELS:
        cr     = sl / 100.0
        order  = _q_at(cr, p50, p85, p95)
        result = _sim(demand, order, cu, co)
        points.append({
            "service_level": sl,
            "order_quantile": f"~P{sl}",
            **result,
        })

    return {"date": date, "gm_placeholder": GM, "points": points}
