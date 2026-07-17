"""
Excel template builder for manual-input data gaps.

build_template(key) -> bytes (xlsx)

Supported keys:
  product_economics  — seed rows from dim_product (ProductId + ProductName),
                       blank columns for gm, cost, shelf_life_days, salvage_frac
  inventory_daily    — seed rows from dim_branch (BranchId + BranchName),
                       blank columns: date, outlet_id, product_id,
                       open_stock, close_stock, received, wasted

Header row: dark-blue (#1E40AF) fill, white bold font.
"""
from __future__ import annotations
import io, pathlib
import duckdb
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent.parent
RAW  = ROOT / "data" / "raw"

_HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col)].width = 22


def _duck_query(sql: str) -> list[dict]:
    c = duckdb.connect(":memory:")
    cur = c.execute(sql)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


def _build_product_economics() -> bytes:
    parquet = (RAW / "dim_product.parquet").as_posix()
    rows = _duck_query(
        f"SELECT ProductId, ProductName FROM read_parquet('{parquet}') "
        f"WHERE CatLvl1_Name = 'FG' ORDER BY ProductId"
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "product_economics"

    headers = ["ProductId", "ProductName", "gm", "cost", "shelf_life_days", "salvage_frac"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in rows:
        ws.append([r["ProductId"], r["ProductName"], None, None, None, None])

    ws.freeze_panes = "C2"  # freeze id+name columns + header

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_inventory_daily() -> bytes:
    parquet = (RAW / "dim_branch.parquet").as_posix()
    rows = _duck_query(
        f"SELECT BranchId, BranchName FROM read_parquet('{parquet}') ORDER BY BranchId"
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "inventory_daily"

    headers = [
        "date", "outlet_id", "product_id",
        "open_stock", "close_stock", "received", "wasted",
        "BranchId_ref", "BranchName_ref",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    # Seed reference rows so user can copy outlet_id values
    for r in rows:
        ws.append([None, None, None, None, None, None, None,
                   r["BranchId"], r["BranchName"]])

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_promo_calendar() -> bytes:
    """Header-only template + an outlet reference sheet (ids to copy)."""
    parquet = (RAW / "dim_branch.parquet").as_posix()
    rows = _duck_query(f"SELECT BranchId, BranchName FROM read_parquet('{parquet}') ORDER BY BranchId")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "promo_calendar"
    headers = ["date", "outlet_id", "product_id", "promo_type", "discount_pct"]
    ws.append(headers)
    _style_header(ws, len(headers))
    # one illustrative blank example row (keeps format obvious)
    ws.append(["2026-08-01", None, None, "%off", None])
    ws.freeze_panes = "A2"

    ref = wb.create_sheet("outlet_ref")
    ref.append(["BranchId", "BranchName"])
    _style_header(ref, 2)
    for r in rows:
        ref.append([r["BranchId"], r["BranchName"]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_lead_time_sla() -> bytes:
    """product_id × warehouse_id -> lead_days, with a warehouse reference sheet."""
    wh = (RAW / "dim_warehouse.parquet").as_posix()
    rows = _duck_query(f"SELECT WarehouseId, WarehouseName FROM read_parquet('{wh}') ORDER BY WarehouseId")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "lead_time_sla"
    headers = ["product_id", "warehouse_id", "lead_days"]
    ws.append(headers)
    _style_header(ws, len(headers))
    ws.freeze_panes = "A2"

    ref = wb.create_sheet("warehouse_ref")
    ref.append(["WarehouseId", "WarehouseName"])
    _style_header(ref, 2)
    for r in rows:
        ref.append([r["WarehouseId"], r["WarehouseName"]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_BUILDERS = {
    "product_economics": _build_product_economics,
    "inventory_daily":   _build_inventory_daily,
    "promo_calendar":    _build_promo_calendar,
    "lead_time_sla":     _build_lead_time_sla,
}


def build_template(key: str) -> bytes:
    """Return xlsx bytes for the requested template key.

    Raises KeyError if key is unknown.
    """
    fn = _BUILDERS.get(key)
    if fn is None:
        raise KeyError(f"No template for key '{key}'. Known: {list(_BUILDERS)}")
    return fn()
